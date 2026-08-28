# rota_geral.py

from __future__ import annotations

import unicodedata
from io import BytesIO, StringIO
from typing import TYPE_CHECKING, Any, Literal, TypedDict, cast

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.utils import get_column_letter

from components.componentes import (
    COR_PRIMARIA,
    COR_SECUNDARIA,
    COR_TEXTO,
    COR_TEXTO_2,
    COR_TEXTO_3,
    FONTE_TEXTO,
    FONTE_TITULO,
    aplicar_estilo,
    render_hero,
    render_insight,
    render_kpi,
    render_kpi_sm,
    render_section_header,
)

# Fallback se COR_LARANJA_SUAVE não existir no componentes
try:
    from components.componentes import COR_LARANJA_SUAVE  # type: ignore
except ImportError:
    COR_LARANJA_SUAVE = "#FFB86B"

if TYPE_CHECKING:
    from streamlit.runtime.uploaded_file_manager import UploadedFile


# ============================================================
# TIPOS
# ============================================================


class LinhaResultado(TypedDict, total=False):
    BASE: str
    WO: int
    OS: int
    ND: int
    RC: int
    MESH: int
    MIGRAÇÃO: int
    GPON: int
    PME: int
    Rotas: int
    Montados: int
    Media_OS: float
    Media_Montados: float


TipoInsight = Literal["ok", "alerta", "critico", "info"]


# ============================================================
# CONFIGURAÇÕES
# ============================================================


class Config:
    """Configurações globais da página e das regras de negócio."""

    PAGE_TITLE: str = "Rota Geral | Totale"
    PAGE_ICON: str = "📊"
    LAYOUT: Literal["centered", "wide"] = "wide"

    BASES: list[str] = ["ABCDM", "GUARULHOS", "LESTE"]

    COLUNAS_INDICADORES: list[str] = [
        "WO",
        "OS",
        "ND",
        "RC",
        "MESH",
        "MIGRAÇÃO",
        "GPON",
        "PME",
    ]

    COLUNAS_METRICAS: list[str] = [
        "Rotas",
        "Média OS",
        "Montados",
        "Média Montados",
    ]

    ORDEM_COLUNAS: list[str] = ["BASE"] + COLUNAS_INDICADORES + COLUNAS_METRICAS

    BASES_CONFIG: dict[str, dict[str, str]] = {
        "ABCDM": {"cor": COR_PRIMARIA},
        "GUARULHOS": {"cor": COR_SECUNDARIA},
        "LESTE": {"cor": COR_LARANJA_SUAVE},
    }

    COLUNAS_ESPERADAS: list[str] = [
        "Contrato",
        "Total de tarefas",
        "Tipo O.S 1",
        "Tipo de Atividade.1",
        "Habilidade de Trabalho",
        "Login do Técnico",
    ]


# ============================================================
# UTILITÁRIOS
# ============================================================


class Utils:
    """Ferramentas utilitárias para tratamento e formatação de dados."""

    @staticmethod
    def normalizar_texto(texto: Any) -> str:
        """Remove acentos, espaços extras e converte para maiúsculas."""
        if pd.isna(texto) or texto is None:
            return ""
        txt = str(texto).strip().upper()
        txt = unicodedata.normalize("NFKD", txt)
        return "".join(c for c in txt if not unicodedata.combining(c))

    @staticmethod
    def preparar_colunas(df: pd.DataFrame) -> pd.DataFrame:
        """Remove espaços no início/fim dos nomes das colunas."""
        df = df.copy()
        df.columns = pd.Index([str(c).strip() for c in df.columns])
        return df

    @staticmethod
    def localizar_coluna(df: pd.DataFrame, nome: str) -> str | None:
        """Busca uma coluna no DataFrame ignorando case e acentos."""
        procurado = Utils.normalizar_texto(nome)
        for coluna in df.columns:
            if Utils.normalizar_texto(coluna) == procurado:
                return str(coluna)
        return None

    @staticmethod
    def localizar_colunas_tipo_os(df: pd.DataFrame) -> list[str]:
        """Localiza dinamicamente as colunas de 'Tipo O.S 1' até 'Tipo O.S 10'."""
        encontradas: list[str] = []
        for i in range(1, 11):
            nomes_possiveis = (
                f"Tipo O.S {i}",
                f"Tipo O.S.{i}",
                f"Tipo OS {i}",
                f"Tipo O.S{i}",
            )
            for nome in nomes_possiveis:
                col = Utils.localizar_coluna(df, nome)
                if col is not None:
                    encontradas.append(col)
                    break
        return encontradas

    @staticmethod
    def to_float(v: Any, default: float = 0.0) -> float:
        """Converte um valor genérico para float de forma segura."""
        if pd.isna(v) or v is None:
            return default
        if isinstance(v, (int, float, bool)):
            return float(v)
        try:
            return float(str(v).replace(",", "."))
        except ValueError:
            return default

    @staticmethod
    def to_int(v: Any, default: int = 0) -> int:
        """Converte um valor genérico para inteiro de forma segura."""
        return int(Utils.to_float(v, float(default)))

    @staticmethod
    def fmt_int(v: Any) -> str:
        """Formata um número inteiro com separador de milhar."""
        return f"{Utils.to_int(v):,}".replace(",", ".")

    @staticmethod
    def fmt_float(v: Any, casas: int = 2) -> str:
        """Formata um número float com separador de decimal."""
        return f"{Utils.to_float(v):.{casas}f}".replace(".", ",")


# ============================================================
# PROCESSAMENTO DE DADOS
# ============================================================


class DataProcessor:
    """Classe responsável pelas regras de negócio e cálculo dos KPIs."""

    @staticmethod
    def criar_flag_gpon(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        coluna = Utils.localizar_coluna(df, "Habilidade de Trabalho")
        if coluna is None:
            df["GPON_FLAG"] = 0
            return df

        serie = df[coluna].fillna("").astype(str).map(Utils.normalizar_texto)
        df["GPON_FLAG"] = serie.str.contains("PON", regex=False).astype(int)
        return df

    @staticmethod
    def carregar_arquivo(arquivo: UploadedFile | None) -> pd.DataFrame | None:
        if arquivo is None:
            return None

        nome = arquivo.name.lower()
        try:
            if nome.endswith(".csv"):
                df = DataProcessor._ler_csv(arquivo)
            elif nome.endswith((".xlsx", ".xls")):
                df = DataProcessor._ler_xlsx(arquivo)
            else:
                raise ValueError("Formato não suportado.")

            df = Utils.preparar_colunas(df)
            return DataProcessor.criar_flag_gpon(df)
        except Exception as erro:
            st.error(f"Erro ao carregar {arquivo.name}: {erro}")
            return None

    @staticmethod
    def _ler_csv(arquivo: UploadedFile) -> pd.DataFrame:
        # Usamos uma lista de tuplas (separador, encoding) em vez de dicionários
        tentativas: list[tuple[str, str]] = [
            (";", "utf-8"),
            (";", "latin-1"),
            (",", "utf-8"),
            (",", "latin-1"),
        ]

        for sep, encoding in tentativas:
            try:
                arquivo.seek(0)
                # Passando explicitamente sep= e encoding= evita o erro do Pylance
                df = pd.read_csv(arquivo, sep=sep, encoding=encoding, low_memory=False)
                if len(df.columns) > 1:
                    return df
            except Exception:
                continue

        raise ValueError("Não foi possível identificar o formato/encoding do CSV.")

    @staticmethod
    def _ler_xlsx(arquivo: UploadedFile) -> pd.DataFrame:
        arquivo.seek(0)
        conteudo_bytes = arquivo.read()

        # Tenta ler como Excel real
        try:
            return pd.read_excel(BytesIO(conteudo_bytes), engine="openpyxl")
        except Exception:
            pass

        # Fallback caso seja um arquivo HTML disfarçado de .xls
        try:
            texto = conteudo_bytes.decode("utf-8")
        except UnicodeDecodeError:
            texto = conteudo_bytes.decode("latin-1")
        return pd.read_html(StringIO(texto))[0]

    @staticmethod
    def _contratos_unicos(df: pd.DataFrame, mascara: pd.Series | None = None) -> int:
        col = Utils.localizar_coluna(df, "Contrato")
        if col is None:
            return 0

        dados = df.loc[mascara] if mascara is not None else df
        s = dados[col].dropna().astype(str).str.strip()
        return int(s[~s.isin(["", "nan", "None", "0"])].nunique())

    @staticmethod
    def _total_tarefas(df: pd.DataFrame, mascara: pd.Series | None = None) -> int:
        col = Utils.localizar_coluna(df, "Total de tarefas")
        if col is None:
            return 0

        dados = df.loc[mascara] if mascara is not None else df
        return int(pd.to_numeric(dados[col], errors="coerce").fillna(0).sum())

    @staticmethod
    def _coluna_contem(df: pd.DataFrame, nome_coluna: str, texto: str) -> pd.Series:
        coluna = Utils.localizar_coluna(df, nome_coluna)
        if coluna is None:
            return pd.Series(False, index=df.index)

        procurado = Utils.normalizar_texto(texto)
        serie = df[coluna].fillna("").astype(str).map(Utils.normalizar_texto)
        return serie.str.contains(procurado, regex=False)

    @staticmethod
    def _tipo_os_contem(df: pd.DataFrame, texto: str) -> pd.Series:
        procurado = Utils.normalizar_texto(texto)
        mascara = pd.Series(False, index=df.index)
        for coluna in Utils.localizar_colunas_tipo_os(df):
            serie = df[coluna].fillna("").astype(str).map(Utils.normalizar_texto)
            mascara |= serie.str.contains(procurado, regex=False)
        return mascara

    @staticmethod
    def calcular_indicadores(df: pd.DataFrame) -> dict[str, int]:
        """Calcula todos os KPIs principais para uma base."""
        mask_adesao = DataProcessor._coluna_contem(df, "Tipo O.S 1", "ADESAO")
        mask_pme = DataProcessor._coluna_contem(df, "Habilidade de Trabalho", "PME")
        mask_gpon = df["GPON_FLAG"] == 1
        mask_pacote = DataProcessor._tipo_os_contem(df, "PACOTE")

        col_ativ = Utils.localizar_coluna(df, "Tipo de Atividade.1")
        if col_ativ:
            mask_rc = df[col_ativ].fillna("").astype(str).map(
                Utils.normalizar_texto
            ) == Utils.normalizar_texto("Retorno Credenciada")
        else:
            mask_rc = pd.Series(False, index=df.index)

        mask_mesh = DataProcessor._tipo_os_contem(df, "MESH")

        return {
            "WO": DataProcessor._contratos_unicos(df),
            "OS": DataProcessor._total_tarefas(df),
            "ND": DataProcessor._contratos_unicos(df, mask_adesao),
            "RC": DataProcessor._contratos_unicos(df, mask_rc),
            "MESH": DataProcessor._total_tarefas(df, mask_mesh),
            "MIGRAÇÃO": DataProcessor._contratos_unicos(df, mask_pacote & mask_gpon),
            "GPON": DataProcessor._contratos_unicos(df, mask_gpon),
            "PME": DataProcessor._contratos_unicos(df, mask_adesao & mask_pme),
        }

    @staticmethod
    def calcular_rotas(df: pd.DataFrame) -> int:
        coluna = Utils.localizar_coluna(df, "Login do Técnico")
        if coluna is None:
            return 0
        logins = df[coluna].dropna().astype(str).str.strip()
        return int(logins[~logins.isin(["", "nan", "None"])].nunique())

    @staticmethod
    def processar_base(nome: str, df: pd.DataFrame, montados: int) -> dict[str, Any]:
        """Gera a linha de resultados consolidada para uma base específica."""
        ind = DataProcessor.calcular_indicadores(df)
        rotas = DataProcessor.calcular_rotas(df)
        os_qtd = int(ind["OS"])

        linha: dict[str, Any] = {"BASE": nome, **ind}
        linha["Rotas"] = rotas
        linha["Média OS"] = round(os_qtd / rotas, 2) if rotas > 0 else 0.0
        linha["Montados"] = int(montados)
        linha["Média Montados"] = round(os_qtd / montados, 2) if montados > 0 else 0.0
        return linha

    @staticmethod
    def adicionar_linha_total(df: pd.DataFrame) -> pd.DataFrame:
        """Adiciona a linha de TOTAL no final do DataFrame consolidado."""
        soma_os = df["OS"].sum()
        soma_rotas = df["Rotas"].sum()
        soma_montados = df["Montados"].sum()

        total: dict[str, Any] = {"BASE": "Total"}
        for col in Config.COLUNAS_INDICADORES + ["Rotas", "Montados"]:
            total[col] = int(df[col].sum())

        total["Média OS"] = round(soma_os / soma_rotas, 2) if soma_rotas > 0 else 0.0
        total["Média Montados"] = (
            round(soma_os / soma_montados, 2) if soma_montados > 0 else 0.0
        )

        return pd.concat([df, pd.DataFrame([total])], ignore_index=True)


# ============================================================
# VISUALIZAÇÃO
# ============================================================


class Visualization:
    """Renderização de Tabelas HTML corporativas e Gráficos Plotly."""

    COLS_NEGATIVAS = {"RC"}  # Menor é melhor

    @staticmethod
    def gerar_excel(df: pd.DataFrame) -> bytes:
        """Gera arquivo Excel formatado para download."""
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Rota Geral", index=False)
            ws = writer.book["Rota Geral"]

            # Estiliza o cabeçalho
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True, color="FFFFFF")
                cell.fill = cell.fill.copy(fill_type="solid", fgColor="012869")

            # Ajusta largura das colunas
            for idx, coluna in enumerate(ws.columns, start=1):
                letra = get_column_letter(idx)
                maior = max(
                    (len(str(c.value)) for c in coluna if c.value is not None),
                    default=0,
                )
                ws.column_dimensions[letra].width = maior + 3

        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _classe_celula(
        valor: float, coluna: str, df_bases: pd.DataFrame | None = None
    ) -> str:
        """Define o estilo (Mapa de Calor) da célula da tabela HTML baseada no ranking."""
        if coluna == "Média OS":
            return "cel-media-os"
        if coluna == "Média Montados":
            return "cel-media-mont"
        if df_bases is None or coluna not in df_bases.columns:
            return "cel-cinza-corp"

        valores = pd.to_numeric(df_bases[coluna], errors="coerce").dropna()
        if len(valores) == 0:
            return "cel-cinza-corp"

        ascending = coluna in Visualization.COLS_NEGATIVAS
        ranking = valores.rank(method="min", ascending=ascending)

        try:
            posicao = int(ranking.loc[valores[valores == valor].index[0]])
            if posicao == 1:
                return "cel-heat-verde"
            if posicao == 2:
                return "cel-heat-amarelo"
            return "cel-heat-vermelho"
        except (IndexError, KeyError):
            return "cel-cinza-corp"

    @staticmethod
    def renderizar_tabela_html(df: pd.DataFrame) -> None:
        """Renderiza a tabela de resultados com layout corporativo HTML/CSS."""
        if df.empty:
            render_insight("Nenhum dado disponível.", "info")
            return

        cols = [str(c) for c in df.columns]
        thead = "".join(f"<th>{c}</th>" for c in cols)
        linhas_html: list[str] = []
        df_bases_ref = df.iloc[:-1].copy()

        for i, row in df.iterrows():
            eh_total = i == len(df) - 1
            classe_tr = ' class="linha-total"' if eh_total else ""

            celulas: list[str] = []
            for c in cols:
                val = row[c]
                if c == "BASE":
                    texto = (
                        str(val)
                        .replace("&", "&amp;")
                        .replace("<", "&lt;")
                        .replace(">", "&gt;")
                    )
                    celulas.append(f'<td class="col-base">{texto}</td>')
                    continue

                v = Utils.to_float(val)
                texto = (
                    Utils.fmt_float(v)
                    if c in ("Média OS", "Média Montados")
                    else Utils.fmt_int(v)
                )

                if eh_total:
                    celulas.append(f'<td class="num">{texto}</td>')
                else:
                    classe = Visualization._classe_celula(v, c, df_bases_ref)
                    celulas.append(f'<td class="num {classe}">{texto}</td>')

            linhas_html.append(f"<tr{classe_tr}>{''.join(celulas)}</tr>")

        tabela = f"""
        <div class="tabela-rota-wrapper">
            <div class="tabela-rota-card">
                <table class="corp-table tabela-rota">
                    <thead><tr>{thead}</tr></thead>
                    <tbody>{"".join(linhas_html)}</tbody>
                </table>
            </div>
        </div>
        """
        st.markdown(tabela, unsafe_allow_html=True)

    @staticmethod
    def _layout_corp(fig: go.Figure, titulo: str = "", height: int = 380) -> go.Figure:
        """Aplica o tema corporativo da Totale nos gráficos do Plotly."""
        fig.update_layout(
            title=(
                dict(
                    text=f"<b>{titulo}</b>",
                    font=dict(family=FONTE_TITULO, size=16, color=COR_PRIMARIA),
                    x=0.01,
                    xanchor="left",
                )
                if titulo
                else fig.layout
            ),
            font=dict(family=FONTE_TEXTO, size=12, color=COR_TEXTO_2),
            height=height,
            margin=dict(t=60, b=40, l=20, r=20),
            paper_bgcolor="white",
            plot_bgcolor="rgba(0,0,0,0)",
            legend=dict(font=dict(family=FONTE_TEXTO, size=11)),
        )
        return fig

    @staticmethod
    def grafico_comparativo_medias(df_bases: pd.DataFrame) -> go.Figure:
        fig = go.Figure()
        fig.add_trace(
            go.Bar(
                name="Média OS",
                x=df_bases["BASE"],
                y=df_bases["Média OS"],
                marker_color=COR_PRIMARIA,
                text=[Utils.fmt_float(v) for v in df_bases["Média OS"]],
                textposition="outside",
                textfont=dict(family=FONTE_TEXTO, size=12),
            )
        )
        fig.add_trace(
            go.Bar(
                name="Média Montados",
                x=df_bases["BASE"],
                y=df_bases["Média Montados"],
                marker_color=COR_SECUNDARIA,
                text=[Utils.fmt_float(v) for v in df_bases["Média Montados"]],
                textposition="outside",
                textfont=dict(family=FONTE_TEXTO, size=12),
            )
        )
        fig.update_layout(barmode="group")
        return Visualization._layout_corp(
            fig, "Média de OS por Rota × por Montado", height=380
        )

    @staticmethod
    def grafico_participacao(df_bases: pd.DataFrame) -> go.Figure:
        color_map = {k: v["cor"] for k, v in Config.BASES_CONFIG.items()}
        fig = px.pie(
            df_bases,
            values="OS",
            names="BASE",
            title="Participação de cada base no total de OS",
            hole=0.6,
            color="BASE",
            color_discrete_map=color_map,
        )
        fig.update_traces(
            textinfo="percent+label",
            textfont=dict(family=FONTE_TEXTO, size=13),
            marker=dict(line=dict(color="white", width=3)),
        )
        fig.update_layout(showlegend=False)
        return Visualization._layout_corp(
            fig, "Participação de cada base no total de OS", height=380
        )

    @staticmethod
    def grafico_heatmap(df_bases: pd.DataFrame) -> go.Figure:
        matriz = df_bases.set_index("BASE")[Config.COLUNAS_INDICADORES]
        matriz_max = matriz.max().replace(0, 1)
        matriz_norm = matriz.div(matriz_max, axis=1)

        escala_totale = [
            [0.00, "#DBEAFE"],
            [0.20, "#BFDBFE"],
            [0.40, "#E0E7FF"],
            [0.55, "#FFF7ED"],
            [0.70, "#FED7AA"],
            [0.85, "#FDBA74"],
            [1.00, "#F37C04"],
        ]

        fig = go.Figure(
            data=go.Heatmap(
                z=matriz_norm.to_numpy(),
                x=list(matriz_norm.columns),
                y=list(matriz_norm.index),
                text=matriz.to_numpy(),
                texttemplate="<b>%{text}</b>",
                textfont=dict(family=FONTE_TEXTO, size=13, color=COR_TEXTO),
                colorscale=escala_totale,
                showscale=True,  # type: ignore
                colorbar=dict(
                    title=dict(
                        text="Intensidade",
                        font=dict(family=FONTE_TEXTO, size=11, color=COR_TEXTO_3),
                    ),
                    thickness=12,
                    len=0.7,
                ),
                hovertemplate="<b>%{y}</b><br>Indicador: <b>%{x}</b><br>Valor: <b>%{text}</b><br>Intensidade: %{z:.0%}<extra></extra>",
                xgap=4,
                ygap=4,
            )
        )
        fig.update_layout(
            xaxis=dict(
                side="bottom",
                tickfont=dict(family=FONTE_TEXTO, size=12, color=COR_TEXTO_2),
                showgrid=False,
                zeroline=False,
            ),
            yaxis=dict(
                tickfont=dict(family=FONTE_TITULO, size=12, color=COR_PRIMARIA),
                showgrid=False,
                zeroline=False,
                autorange="reversed",
            ),
        )
        return Visualization._layout_corp(
            fig, "Intensidade relativa por indicador", height=340
        )

    @staticmethod
    def gerar_insights(
        df_bases: pd.DataFrame, total: dict[str, Any]
    ) -> list[tuple[str, TipoInsight]]:
        """Gera análises textuais automáticas baseadas nos resultados."""
        insights: list[tuple[str, TipoInsight]] = []

        top_os = df_bases.loc[df_bases["Média OS"].idxmax()]
        insights.append(
            (
                f"🥇 **{top_os['BASE']}** possui a maior Média OS: **{Utils.fmt_float(top_os['Média OS'])}** OS por rota.",
                "ok",
            )
        )

        low_os = df_bases.loc[df_bases["Média OS"].idxmin()]
        if str(low_os["BASE"]) != str(top_os["BASE"]):
            insights.append(
                (
                    f"⬇️ **{low_os['BASE']}** apresenta a menor Média OS: **{Utils.fmt_float(low_os['Média OS'])}** — avaliar produtividade.",
                    "alerta",
                )
            )

        base_max = df_bases.loc[df_bases["OS"].idxmax()]
        total_os = Utils.to_float(total.get("OS"))
        valor_base = Utils.to_float(base_max["OS"])
        perc = (valor_base / total_os * 100) if total_os > 0 else 0.0

        insights.append(
            (
                f"📊 **{base_max['BASE']}** concentra **{Utils.fmt_float(perc, 1)}%** do total de OS ({Utils.fmt_int(valor_base)} de {Utils.fmt_int(total_os)}).",
                "info",
            )
        )

        total_mont = Utils.to_float(total.get("Montados"))
        if total_mont > 0:
            aprov = total_os / total_mont
            tipo: TipoInsight = (
                "ok" if aprov >= 8 else "alerta" if aprov >= 5 else "critico"
            )
            insights.append(
                (
                    f"🎯 Aproveitamento geral: **{Utils.fmt_float(aprov)}** OS por montado ({Utils.fmt_int(total_mont)} montados no total).",
                    tipo,
                )
            )

        for col in Config.COLUNAS_INDICADORES:
            zerados = df_bases[df_bases[col] == 0]["BASE"].astype(str).tolist()
            if zerados and len(zerados) < len(df_bases):
                insights.append(
                    (
                        f"⚠️ Indicador **{col}** está zerado em: "
                        + ", ".join(f"**{b}**" for b in zerados),
                        "alerta",
                    )
                )

        return insights


# ============================================================
# INTERFACE (Streamlit UI)
# ============================================================


class UI:
    """Responsável por toda a montagem visual da página no Streamlit."""

    @staticmethod
    def reiniciar_painel() -> None:
        """Limpa o cache da sessão e reseta a página."""
        for k in list(st.session_state.keys()):
            if str(k).startswith(("up_", "mont_", "btn_")):
                del st.session_state[k]
        st.session_state["_reset_counter"] = (
            int(st.session_state.get("_reset_counter", 0)) + 1
        )
        st.rerun()

    @staticmethod
    def validar_colunas(nome_base: str, df: pd.DataFrame) -> None:
        """Alerta o usuário caso faltem colunas obrigatóres na base enviada."""
        faltantes = [
            c for c in Config.COLUNAS_ESPERADAS if Utils.localizar_coluna(df, c) is None
        ]
        if faltantes:
            render_insight(
                f"**{nome_base}**: colunas não encontradas: "
                + ", ".join(f"`{c}`" for c in faltantes),
                "alerta",
            )

    @staticmethod
    def mostrar_interface() -> None:
        """Função principal que desenha a tela."""
        st.set_page_config(
            page_title=Config.PAGE_TITLE,
            page_icon=Config.PAGE_ICON,
            layout=Config.LAYOUT,
        )
        aplicar_estilo()
        render_hero(
            titulo="📊 Rota Geral | TOTALE",
            subtitulo="Ponto de partida da roteirização, unificando os dados das bases operacionais",
            badge="Operacional",
        )
        UI._injetar_css()

        render_section_header(
            icon="📁", title="Importação das bases", badge="CSV ou XLSX"
        )
        arquivos = UI._mostrar_uploaders()

        render_section_header(
            icon="📝", title="Montados (digitação manual)", badge="Entrada"
        )
        montados_dict = UI._mostrar_inputs_montados()

        UI._mostrar_botoes_acao(arquivos, montados_dict)

    @staticmethod
    def _injetar_css() -> None:
        """Injeta a folha de estilos CSS personalizada no Streamlit com suporte a design tokens."""
        st.markdown(
            f"""
            <style>
            /* ==========================================================================
               01. DESIGN TOKENS / VARIÁVEIS CSS
               ========================================================================== */
            :root {{
                /* Tipografia */
                --font-titulo: {FONTE_TITULO}, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
                --font-texto: {FONTE_TEXTO}, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;

                /* Paleta Corporativa (Navy & Laranja) */
                --brand-navy-dark: #012869;
                --brand-navy-mid: #023A9E;
                --brand-navy-light: #1E5FCC;
                
                --brand-orange-light: #F37C04;
                --brand-orange-mid: #E85D04;
                --brand-orange-dark: #C44100;

                /* Gradientes Principais */
                --grad-primary: linear-gradient(135deg, var(--brand-navy-dark) 0%, var(--brand-navy-mid) 40%, var(--brand-orange-light) 100%);
                --grad-navy: linear-gradient(135deg, var(--brand-navy-dark) 0%, var(--brand-navy-mid) 55%, var(--brand-navy-light) 100%);
                --grad-orange: linear-gradient(135deg, var(--brand-orange-light) 0%, var(--brand-orange-mid) 55%, var(--brand-orange-dark) 100%);
                --grad-success: linear-gradient(135deg, #059669 0%, #10B981 100%);
                --grad-neutral: linear-gradient(135deg, #F3F4F6 0%, #E5E7EB 100%);
                --grad-total-row: linear-gradient(135deg, #1F2937 0%, #374151 45%, #4B5563 80%, #6B7280 100%);

                /* Heatmap & Indicadores */
                --heat-green-bg: linear-gradient(180deg, #D1FAE5 0%, #A7F3D0 100%);
                --heat-green-text: #065F46;
                --heat-green-border: #10B981;

                --heat-yellow-bg: linear-gradient(180deg, #FEF3C7 0%, #FDE68A 100%);
                --heat-yellow-text: #92400E;
                --heat-yellow-border: #F59E0B;

                --heat-red-bg: linear-gradient(180deg, #FEE2E2 0%, #FECACA 100%);
                --heat-red-text: #991B1B;
                --heat-red-border: #EF4444;

                /* Bordas, Sombras e Transições */
                --border-light: #E2E8F0;
                --border-radius-card: 16px;
                --border-radius-btn: 10px;
                --shadow-card: 0 10px 30px rgba(0, 0, 0, 0.08), 0 2px 6px rgba(0, 0, 0, 0.04);
                --transition-base: all 0.25s cubic-bezier(0.4, 0, 0.2, 1);
            }}

            /* ==========================================================================
               02. LAYOUT GERAL
               ========================================================================== */
            .block-container {{
                padding-top: 1.5rem !important;
                padding-bottom: 3rem !important;
            }}

            /* ==========================================================================
               03. COMPONENTES STREAMLIT (Inputs & Uploaders)
               ========================================================================== */
            [data-testid="stFileUploader"] {{
                background: linear-gradient(180deg, #FFFFFF 0%, #F9FAFB 100%);
                border: 2px dashed #CBD5E1;
                border-radius: 12px;
                padding: 10px;
                transition: var(--transition-base);
            }}
            [data-testid="stFileUploader"]:hover {{
                border-color: var(--brand-navy-dark);
                box-shadow: 0 4px 12px rgba(1, 40, 105, 0.08);
            }}
            [data-testid="stFileUploader"] button {{
                background: linear-gradient(135deg, var(--brand-navy-dark) 0%, var(--brand-navy-mid) 100%) !important;
                color: #FFFFFF !important;
                border: none !important;
                border-radius: 6px !important;
                font-weight: 600 !important;
            }}

            [data-testid="stNumberInput"] input {{
                border-radius: 8px !important;
                border: 1.5px solid var(--border-light) !important;
                font-weight: 600 !important;
                font-size: 15px !important;
                color: var(--brand-navy-dark) !important;
            }}

            /* Botões */
            .stButton > button[kind="primary"] {{
                background: var(--grad-primary) !important;
                color: #FFFFFF !important;
                border: none !important;
                padding: 14px 28px !important;
                border-radius: var(--border-radius-btn) !important;
                font-weight: 700 !important;
                font-size: 15px !important;
                box-shadow: 0 4px 14px rgba(1, 40, 105, 0.30) !important;
                transition: var(--transition-base);
            }}
            .stButton > button[kind="primary"]:hover {{
                opacity: 0.95;
                transform: translateY(-1px);
            }}

            .stDownloadButton > button {{
                background: var(--grad-success) !important;
                color: #FFFFFF !important;
                border: none !important;
                padding: 12px 20px !important;
                border-radius: var(--border-radius-btn) !important;
                font-weight: 700 !important;
                transition: var(--transition-base);
            }}

            .stButton > button[kind="secondary"] {{
                background: var(--grad-neutral) !important;
                color: #374151 !important;
                border: 1.5px solid #D1D5DB !important;
                padding: 14px 28px !important;
                border-radius: var(--border-radius-btn) !important;
                font-weight: 700 !important;
                transition: var(--transition-base);
            }}
            
            /* ==========================================================================
            04. TABELA DE ROTA  (centralizada + auto-ajuste)
            ========================================================================== */

            /* Wrapper: centraliza a tabela na página */
            .tabela-rota-wrapper {{
                display: flex !important;
                justify-content: center !important;
                width: 100% !important;
                margin: 20px 0 32px 0 !important;
                padding: 0 !important;
                overflow-x: auto !important;          /* scroll só se a tela for estreita */
                -webkit-overflow-scrolling: touch;
                background: transparent !important;
                border: none !important;
                box-shadow: none !important;
            }}

            /* Card interno: abraça o conteúdo e centraliza */
            .tabela-rota-wrapper > .tabela-rota-card,
            .tabela-rota-wrapper table.tabela-rota {{
                margin-left: auto !important;
                margin-right: auto !important;
            }}

            /* Card visual da tabela */
            .tabela-rota-card {{
                display: inline-block !important;
                width: fit-content !important;
                max-width: 100% !important;
                border-radius: var(--border-radius-card) !important;
                box-shadow: var(--shadow-card) !important;
                border: 1px solid var(--border-light) !important;
                background: #FFFFFF !important;
                overflow: hidden !important;          /* cantos arredondados limpos */
            }}

            /* Tabela: largura pelo conteúdo, sem % rígidos */
            .tabela-rota {{
                border-collapse: separate !important;
                border-spacing: 0 !important;
                width: max-content !important;        /* cresce conforme as colunas */
                max-width: 100% !important;
                table-layout: auto !important;        /* auto-ajuste real */
                margin: 0 auto !important;
            }}

            /* Colunas: padding fluido, sem min-width agressivo */
            .tabela-rota th,
            .tabela-rota td {{
                width: auto !important;
                min-width: 0 !important;
                padding: 14px 12px !important;        /* um pouco mais de ar horizontal */
                box-sizing: border-box !important;
            }}

            /* Primeira coluna (BASE) um pouco mais larga por legibilidade */
            .tabela-rota th:first-child,
            .tabela-rota td:first-child {{
                padding-left: 18px !important;
                padding-right: 18px !important;
                white-space: nowrap !important;
            }}

            /* Cabeçalho */
            .tabela-rota thead th {{
                background: var(--grad-orange) !important;
                color: #FFFFFF !important;
                font-family: var(--font-titulo) !important;
                font-weight: 700 !important;
                text-transform: uppercase !important;
                letter-spacing: 0.7px !important;
                font-size: 11px !important;
                padding: 16px 12px !important;
                text-align: center !important;
                border-bottom: none !important;
                text-shadow: 0 1px 2px rgba(0,0,0,0.25);
                white-space: nowrap !important;
            }}

            .tabela-rota thead th:first-child {{ border-top-left-radius: 14px; }}
            .tabela-rota thead th:last-child  {{ border-top-right-radius: 14px; }}

            /* Corpo */
            .tabela-rota tbody td {{
                font-family: var(--font-texto) !important;
                text-align: center !important;
                border-bottom: 1px solid #F1F5F9 !important;
                font-variant-numeric: tabular-nums !important;
                font-weight: 600 !important;
                font-size: 13px !important;
                white-space: nowrap !important;
                transition: filter 0.15s ease;
            }}

            .tabela-rota tbody tr:hover td {{
                filter: brightness(1.04) saturate(1.06);
            }}

            /* Coluna BASE */
            .tabela-rota td.col-base {{
                background: var(--grad-navy) !important;
                color: #FFFFFF !important;
                font-family: var(--font-titulo) !important;
                font-weight: 800 !important;
                letter-spacing: 0.8px;
                text-transform: uppercase;
                font-size: 12.5px !important;
                text-shadow: 0 1px 2px rgba(0,0,0,0.30);
            }}

            /* Linha TOTAL */
            .tabela-rota tbody tr.linha-total td {{
                background: var(--grad-total-row) !important;
                color: #FFFFFF !important;
                font-weight: 800 !important;
                font-size: 13.5px !important;
                border-bottom: none !important;
                padding: 18px 12px !important;
            }}

            .tabela-rota tbody tr.linha-total td.col-base {{
                background: linear-gradient(135deg, #0F172A 0%, #1F2937 55%, #334155 100%) !important;
            }}

            /* Heatmap / indicadores (inalterados visualmente) */
            .cel-heat-verde {{
                background: var(--heat-green-bg) !important;
                color: var(--heat-green-text) !important;
                font-weight: 700 !important;
                border-left: 3px solid var(--heat-green-border) !important;
            }}
            .cel-heat-amarelo {{
                background: var(--heat-yellow-bg) !important;
                color: var(--heat-yellow-text) !important;
                font-weight: 700 !important;
                border-left: 3px solid var(--heat-yellow-border) !important;
            }}
            .cel-heat-vermelho {{
                background: var(--heat-red-bg) !important;
                color: var(--heat-red-text) !important;
                font-weight: 700 !important;
                border-left: 3px solid var(--heat-red-border) !important;
            }}
            .cel-cinza-corp {{
                background: var(--grad-neutral) !important;
                color: #374151 !important;
                font-weight: 600 !important;
            }}

            table.tabela-rota tbody td.cel-media-os {{
                background: linear-gradient(180deg, var(--brand-navy-light) 0%, var(--brand-navy-mid) 100%) !important;
                color: #FFFFFF !important;
                font-weight: 800 !important;
                text-shadow: 0 1px 2px rgba(0,0,0,0.4) !important;
            }}
            table.tabela-rota tbody td.cel-media-mont {{
                background: linear-gradient(180deg, var(--brand-orange-light) 0%, var(--brand-orange-dark) 100%) !important;
                color: #FFFFFF !important;
                font-weight: 800 !important;
                text-shadow: 0 1px 2px rgba(0,0,0,0.4) !important;
            }}

            /* Telas estreitas: permite scroll horizontal sem quebrar o centro */
            @media (max-width: 900px) {{
                .tabela-rota-wrapper {{
                    justify-content: flex-start !important;
                }}
                .tabela-rota th,
                .tabela-rota td {{
                    padding: 12px 8px !important;
                    font-size: 12px !important;
                }}
            }}
            </style>
            """,
            unsafe_allow_html=True,
        )

    @staticmethod
    def _mostrar_uploaders() -> dict[str, UploadedFile | None]:
        reset = int(st.session_state.get("_reset_counter", 0))
        cols = st.columns(3)
        return {
            base: cols[i].file_uploader(
                f"Base {base}", type=["csv", "xlsx"], key=f"up_{base}_{reset}"
            )
            for i, base in enumerate(Config.BASES)
        }

    @staticmethod
    def _mostrar_inputs_montados() -> dict[str, int]:
        reset = int(st.session_state.get("_reset_counter", 0))
        cols = st.columns(3)
        return {
            base: int(
                cols[i].number_input(
                    f"Montados {base}",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"mont_{base}_{reset}",
                )
            )
            for i, base in enumerate(Config.BASES)
        }

    @staticmethod
    def _mostrar_botoes_acao(
        arquivos: dict[str, UploadedFile | None], montados_dict: dict[str, int]
    ) -> None:
        st.markdown("<br>", unsafe_allow_html=True)
        col1, col2 = st.columns([3, 1])

        processar = col1.button(
            "🚀 Processar bases",
            type="primary",
            use_container_width=True,
            key="btn_processar",
        )
        if col2.button(
            "🔄 Reiniciar painel",
            type="secondary",
            use_container_width=True,
            key="btn_reiniciar",
        ):
            UI.reiniciar_painel()

        if processar:
            UI._processar_bases(arquivos, montados_dict)

    @staticmethod
    def _processar_bases(
        arquivos: dict[str, UploadedFile | None], montados_dict: dict[str, int]
    ) -> None:
        if not all(arquivos.values()):
            render_insight("Selecione as **três bases** antes de processar.", "critico")
            st.stop()

        with st.spinner("Processando as bases..."):
            bases_dfs = {
                nome: DataProcessor.carregar_arquivo(arq)
                for nome, arq in arquivos.items()
            }
            if any(df is None for df in bases_dfs.values()):
                st.stop()

            # Aqui garantimos ao MyPy/Typing que não há mais None nos DFs
            bases_validas = {
                nome: df for nome, df in bases_dfs.items() if df is not None
            }

            for nome, df in bases_validas.items():
                UI.validar_colunas(nome, df)

            resultado = [
                DataProcessor.processar_base(nome, df, montados_dict[nome])
                for nome, df in bases_validas.items()
            ]
            df_bases = pd.DataFrame(resultado)[Config.ORDEM_COLUNAS]
            df_final = DataProcessor.adicionar_linha_total(df_bases)

        render_insight("Bases processadas com sucesso.", "ok")
        total: dict[str, Any] = {
            str(k): v for k, v in df_final.iloc[-1].to_dict().items()
        }

        UI._mostrar_kpis_gerais(total)
        UI._mostrar_kpis_por_base(df_bases)

        render_section_header(
            icon="📋", title="Resultado consolidado", badge="Detalhamento"
        )
        Visualization.renderizar_tabela_html(df_final)

        st.markdown(
            f"""
            <div style="display:flex;gap:16px;justify-content:flex-end;align-items:center; padding:8px 4px;font-size:12px;font-family:{FONTE_TEXTO}; color:#4B5563;font-weight:600;">
                <span style="display:flex;align-items:center;gap:6px;"><span style="width:14px;height:14px;background:linear-gradient(180deg,#D1FAE5,#A7F3D0); border-left:3px solid #10B981;border-radius:3px;"></span>Melhor</span>
                <span style="display:flex;align-items:center;gap:6px;"><span style="width:14px;height:14px;background:linear-gradient(180deg,#FEF3C7,#FDE68A); border-left:3px solid #F59E0B;border-radius:3px;"></span>Intermediário</span>
                <span style="display:flex;align-items:center;gap:6px;"><span style="width:14px;height:14px;background:linear-gradient(180deg,#FEE2E2,#FECACA); border-left:3px solid #EF4444;border-radius:3px;"></span>Pior</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        UI._mostrar_graficos(df_bases)
        UI._mostrar_insights(df_bases, total)
        UI._mostrar_diagnostico(bases_validas)

        render_section_header(icon="⬇️", title="Exportação", badge="Excel")
        st.download_button(
            "Baixar resultado em Excel",
            data=Visualization.gerar_excel(df_final),
            file_name="rota_geral.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    @staticmethod
    def _mostrar_kpis_gerais(total: dict[str, Any]) -> None:
        render_section_header(
            icon="📈", title="Indicadores gerais", badge="Total consolidado"
        )
        k = st.columns(4)
        render_kpi(
            k[0],
            "Total OS",
            Utils.fmt_int(total.get("OS")),
            "Soma de Total de tarefas",
            "azul",
        )
        render_kpi(
            k[1], "Rotas", Utils.fmt_int(total.get("Rotas")), "Logins únicos", "laranja"
        )
        render_kpi(
            k[2],
            "Média OS",
            Utils.fmt_float(total.get("Média OS")),
            "OS por rota",
            "verde",
        )
        render_kpi(
            k[3],
            "Média Montados",
            Utils.fmt_float(total.get("Média Montados")),
            "OS por montado",
            "vermelho",
        )

    @staticmethod
    def _mostrar_kpis_por_base(df_bases: pd.DataFrame) -> None:
        render_section_header(
            icon="🏢", title="Desempenho por base", badge="Comparativo rápido"
        )
        cols = st.columns(len(df_bases))
        # enumerate garante que 'i' seja sempre um número inteiro seguro para cols[i]
        for i, (_, row) in enumerate(df_bases.iterrows()):
            with cols[i]:
                st.markdown(f"##### 🏷️ {row['BASE']}")
                render_kpi_sm(
                    cast(Any, st),
                    "OS",
                    Utils.fmt_int(row["OS"]),
                    f"{Utils.fmt_int(row['Rotas'])} rotas",
                    "azul",
                )
                render_kpi_sm(
                    cast(Any, st),
                    "Média OS",
                    Utils.fmt_float(row["Média OS"]),
                    "OS por rota",
                    "verde",
                )
                render_kpi_sm(
                    cast(Any, st),
                    "Média Montados",
                    Utils.fmt_float(row["Média Montados"]),
                    "OS por montado",
                    "laranja",
                )

    @staticmethod
    def _mostrar_graficos(df_bases: pd.DataFrame) -> None:
        render_section_header(
            icon="📊", title="Análise visual", badge="Gráficos comparativos"
        )
        g1, g2 = st.columns(2)
        with g1:
            st.plotly_chart(
                Visualization.grafico_participacao(df_bases),
                use_container_width=True,
                config={"displayModeBar": False},
            )
        with g2:
            st.plotly_chart(
                Visualization.grafico_comparativo_medias(df_bases),
                use_container_width=True,
                config={"displayModeBar": False},
            )

        st.plotly_chart(
            Visualization.grafico_heatmap(df_bases),
            use_container_width=True,
            config={"displayModeBar": False},
        )

    @staticmethod
    def _mostrar_insights(df_bases: pd.DataFrame, total: dict[str, Any]) -> None:
        render_section_header(
            icon="💡", title="Insights automáticos", badge="Leitura dos dados"
        )
        for msg, tipo in Visualization.gerar_insights(df_bases, total):
            render_insight(msg, tipo=tipo)

    @staticmethod
    def _mostrar_diagnostico(bases_validas: dict[str, pd.DataFrame]) -> None:
        with st.expander("🔎 Conferência técnica dos cálculos"):
            for nome, df in bases_validas.items():
                st.markdown(f"**{nome}**")
                st.write(f"Linhas na base: {Utils.fmt_int(len(df))}")
                st.write(
                    "Colunas Tipo O.S encontradas:", Utils.localizar_colunas_tipo_os(df)
                )
                st.write("Linhas com GPON_FLAG = 1:", int((df["GPON_FLAG"] == 1).sum()))
                st.divider()


if __name__ == "__main__":
    UI.mostrar_interface()
