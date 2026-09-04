"""
quebra.py
=========
Super Relatório Corporativo Unificado | Quebra Operacional TOTALE

Módulos integrados:
  • Resumo Executivo (Matriz Monitor × Segmento)
  • Análise Detalhada (Projeções, Rankings, Causas, Backoffice, Base)
  • Auditoria de Critérios de Classificação

Critérios centralizados em: components.criterios
"""

from __future__ import annotations

import csv
import sys
from datetime import datetime
from html import escape
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Tuple, cast

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Path bootstrap ──────────────────────────────────────────────────
_DIR = Path(__file__).resolve().parent
_ROOT = _DIR.parent
for _p in (_DIR, _ROOT):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

# ── Componentes visuais globais ─────────────────────────────────────
from components.componentes import (
    FONTE_TEXTO,
    FONTE_TITULO,
    aplicar_estilo,
    render_section_header,
    render_table_html,
)
from components.componentes import render_insight as _render_insight_global
from components.componentes import render_kpi as _render_kpi_global
from components.componentes import render_kpi_sm as _render_kpi_sm_global

# ── Critérios centralizados ─────────────────────────────────────────
from components.criterios import (
    VAZIOS_CONTRATO,
    classificar_tipo_servico,
    detectar_col_contrato,
    detectar_col_status_atividade,
    render_debug_criterios,
    render_painel_criterios,
)

TipoInsight = Literal["ok", "info", "alerta", "critico", "acao"]
TemaKPI = Literal[
    "azul", "verde", "vermelho", "laranja", "cinza", "roxo", "amarelo", "escuro"
]


# ═══════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO DA PÁGINA
# ═══════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Quebra Operacional | TOTALE",
    page_icon="📉",
    layout="wide",
    initial_sidebar_state="expanded",
)
aplicar_estilo()

if "df_memoria" not in st.session_state:
    st.session_state["df_memoria"] = None


# ═══════════════════════════════════════════════════════════════════════
# CONSTANTES DE DOMÍNIO
# ═══════════════════════════════════════════════════════════════════════
class Config:
    """Configurações centralizadas de SLA, cores e domínio operacional."""

    SLA_QUEBRA_MAXIMA = 0.20

    URL_LISTA_ATIVOS = (
        "https://docs.google.com/spreadsheets/d/"
        "1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg/edit"
    )
    SHEET_ID_ATIVOS = "1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg"
    WORKSHEET_ATIVOS = "lista_ativos"

    CONTRATO_VALORES_VAZIOS = VAZIOS_CONTRATO
    STATUS_ORDEM = ["Executada", "Não Executada", "Pendente"]

    CORES_STATUS = {
        "Executada": "#10B981",
        "Não Executada": "#EF4444",
        "Pendente": "#94A3B8",
    }
    COL_REGIAO = "REGIÃO"

    CORES_TIPO = {
        "Novos Domicílios": "#1E40AF",
        "Quebra Geral": "#78350F",
        "Outros": "#64748B",
    }
    ORDEM_TIPOS = ["Novos Domicílios"]


CORES_REGIAO: Dict[str, Dict[str, str]] = {
    "LESTE": {"bg": "#DBEAFE", "text": "#1E40AF", "border": "#3B82F6"},
    "GRU": {"bg": "#D1FAE5", "text": "#065F46", "border": "#10B981"},
    "ABCDM": {"bg": "#EDE9FE", "text": "#5B21B6", "border": "#8B5CF6"},
    "OUTRAS": {"bg": "#F1F5F9", "text": "#475569", "border": "#94A3B8"},
}

TEMAS_CARD_EXTRA: Dict[str, Dict[str, str]] = {
    "amarelo": {
        "fundo": "#FEF9C3",
        "texto": "#854D0E",
        "borda": "#EAB308",
        "titulo": "#A16207",
    },
    "roxo": {
        "fundo": "#FAF5FF",
        "texto": "#7E22CE",
        "borda": "#A855F7",
        "titulo": "#6B21A8",
    },
    "escuro": {
        "fundo": "#1E293B",
        "texto": "#FFFFFF",
        "borda": "#475569",
        "titulo": "#E2E8F0",
    },
}

_MAPA_TEMA_GLOBAL: Dict[str, str] = {
    "azul": "azul",
    "verde": "verde",
    "vermelho": "vermelho",
    "laranja": "laranja",
    "cinza": "cinza",
    "roxo": "azul",
    "amarelo": "laranja",
    "escuro": "cinza",
}

FONTE_CARD_TITULO = '"Poppins", "Manrope", sans-serif'
FONTE_CARD_TEXTO = '"Inter", "Roboto", sans-serif'


# ═══════════════════════════════════════════════════════════════════════
# WRAPPERS DE INTERFACE (UI)
# ═══════════════════════════════════════════════════════════════════════
def render_kpi(
    col: Any, label: str, value: str, sub: str = "", tema: str = "azul"
) -> None:
    if tema in TEMAS_CARD_EXTRA:
        t = TEMAS_CARD_EXTRA[tema]
        col.markdown(
            f'<div style="background:{t["fundo"]};border-left:4px solid {t["borda"]};'
            f'border-radius:10px;padding:20px 24px;box-shadow:0 4px 12px rgba(0,0,0,0.08);">'
            f'<div style="font-family:{FONTE_TEXTO};font-size:11px;font-weight:700;'
            f'color:{t["titulo"]};text-transform:uppercase;letter-spacing:1.2px;'
            f'margin-bottom:6px;">{label}</div>'
            f'<div style="font-family:{FONTE_TITULO};font-size:28px;font-weight:800;'
            f'color:{t["texto"]};line-height:1;font-variant-numeric:tabular-nums;">{value}</div>'
            f'<div style="font-family:{FONTE_TEXTO};font-size:12px;color:{t["titulo"]};'
            f'margin-top:6px;font-weight:500;">{sub}</div></div>',
            unsafe_allow_html=True,
        )
    else:
        _render_kpi_global(col, label, value, sub, _MAPA_TEMA_GLOBAL.get(tema, "azul"))  # type: ignore


def render_kpi_sm(
    col: Any, label: str, value: str, sub: str = "", tema: str = "azul"
) -> None:
    if tema in TEMAS_CARD_EXTRA:
        t = TEMAS_CARD_EXTRA[tema]
        col.markdown(
            f'<div style="background:{t["fundo"]};border-left:3px solid {t["borda"]};'
            f"border-radius:6px;padding:12px 16px;margin-bottom:8px;"
            f'box-shadow:0 1px 4px rgba(0,0,0,0.06);">'
            f'<div style="font-family:{FONTE_TEXTO};font-size:10px;color:{t["titulo"]};'
            f'text-transform:uppercase;letter-spacing:1px;font-weight:700;">{label}</div>'
            f'<div style="font-family:{FONTE_TITULO};font-size:20px;color:{t["texto"]};'
            f"font-weight:800;line-height:1.2;margin-top:4px;"
            f'font-variant-numeric:tabular-nums;">{value}</div>'
            f'<div style="font-family:{FONTE_TEXTO};font-size:11px;color:{t["titulo"]};'
            f'margin-top:2px;">{sub}</div></div>',
            unsafe_allow_html=True,
        )
    else:
        _render_kpi_sm_global(col, label, value, sub, _MAPA_TEMA_GLOBAL.get(tema, "azul"))  # type: ignore


def render_insight(texto: str, tipo: TipoInsight = "info") -> None:
    _render_insight_global(texto, tipo)


def render_section(titulo: str) -> None:
    partes = titulo.strip().split(" ", 1)
    primeiro_char = partes[0][0] if partes[0] else ""
    if len(partes) == 2 and not primeiro_char.isascii():
        icon, title = partes[0], partes[1]
    else:
        icon, title = "📊", titulo
    render_section_header(icon, title)


# ═══════════════════════════════════════════════════════════════════════
# UTILITÁRIOS OPERACIONAIS
# ═══════════════════════════════════════════════════════════════════════
def _fmt_pct_br(v: Any) -> str:
    try:
        val = float(v) * 100
        return f"{val:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "0,00%"


def _fmt_int_br(v: Any) -> str:
    try:
        return f"{int(float(v)):,}".replace(",", ".")
    except (ValueError, TypeError):
        return "0"


class Utils:
    """Utilitários de manipulação, busca de colunas e exportação."""

    @staticmethod
    def buscar_coluna(df: pd.DataFrame, palavras: list) -> Optional[str]:
        if df is None or df.empty:
            return None
        cols = {
            str(c)
            .strip()
            .upper()
            .replace(".", "")
            .replace("_", "")
            .replace("  ", " "): c
            for c in df.columns
        }
        for p in palavras:
            pn = (
                str(p)
                .strip()
                .upper()
                .replace(".", "")
                .replace("_", "")
                .replace("  ", " ")
            )
            for cn, co in cols.items():
                if pn in cn:
                    return co
        return None

    @staticmethod
    def classificar_status(serie: pd.Series) -> pd.Series:
        s = serie.fillna("").astype(str).str.strip().str.upper()
        exe = s == "EXECUTADA"
        nex = s.isin(["NÃO EXECUTADA", "NAO EXECUTADA"])
        return pd.Series(
            np.select([exe, nex], ["Executada", "Não Executada"], default="Pendente"),
            index=serie.index,
        )

    @staticmethod
    def gerar_excel(df: pd.DataFrame, aba: str = "Dados") -> bytes:
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            df.to_excel(w, index=False, sheet_name=aba[:31])
            ws = w.sheets[aba[:31]]

            # Habilitar linhas de grade explicitamente
            ws.views.sheetView[0].showGridLines = True

            # Estilos OpenPyXL
            header_fill = PatternFill("solid", fgColor="0F172A")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            border_thin = Side(border_style="thin", color="CBD5E1")
            cell_border = Border(
                left=border_thin, right=border_thin, top=border_thin, bottom=border_thin
            )
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center

            # Formatando as linhas de dados com tipos corretos de dados no Excel
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = cell_border
                    val = cell.value

                    col_name = str(ws.cell(1, col).value).upper()

                    try:
                        if val is not None and str(val).strip() != "":
                            if (
                                "%" in col_name
                                or "QUEBRA" in col_name
                                or "TAXA" in col_name
                            ):
                                cell.value = (
                                    float(
                                        str(val)
                                        .replace("%", "")
                                        .replace(",", ".")
                                        .strip()
                                    )
                                    / 100.0
                                )
                                cell.number_format = "0.0%"
                            elif col_name in [
                                "VOLUME",
                                "ALOCADO",
                                "EXECUTADA",
                                "NÃO EXECUTADA",
                                "PENDENTE",
                                "TOTAL DE TAREFAS",
                                "TOTAL TAREFAS",
                            ]:
                                cell.value = int(
                                    float(
                                        str(val)
                                        .replace(".", "")
                                        .replace(",", ".")
                                        .strip()
                                    )
                                )
                                cell.number_format = "#,##0"
                            elif any(
                                x in col_name for x in ["PROBAB", "PROJ", "FECHAMENTO"]
                            ):
                                v_flt = float(
                                    str(val).replace("%", "").replace(",", ".").strip()
                                )
                                if v_flt > 1.0:
                                    v_flt = v_flt / 100.0
                                cell.value = v_flt
                                cell.number_format = "0.0%"
                    except Exception:
                        pass

                    if isinstance(cell.value, (int, float)):
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left

            for i, col in enumerate(df.columns, 1):
                try:
                    serie_str = df[col].fillna("").astype(str)
                    tamanhos = serie_str.str.len()
                    max_len_dados = int(tamanhos.max()) if len(tamanhos) > 0 else 0
                    max_len = max(max_len_dados, len(str(col)))
                    ws.column_dimensions[get_column_letter(i)].width = min(
                        max(max_len + 3, 12), 40
                    )
                except Exception:
                    ws.column_dimensions[get_column_letter(i)].width = 20
        return out.getvalue()


# ═══════════════════════════════════════════════════════════════════════
# CARREGAMENTO E SANEAMENTO DE DADOS (ETL)
# ═══════════════════════════════════════════════════════════════════════
class DataLoader:
    """Pipeline de Ingestão, limpeza e enriquecimento da base."""

    @staticmethod
    @st.cache_data(show_spinner=False)
    def ler_arquivo(file_bytes: bytes, filename: str) -> pd.DataFrame:
        bio = BytesIO(file_bytes)
        try:
            if filename.lower().endswith(".csv"):
                bio.seek(0)
                amostra = bio.read(5000).decode("utf-8", errors="ignore")
                bio.seek(0)
                try:
                    sep = csv.Sniffer().sniff(amostra).delimiter if amostra else ";"
                except Exception:
                    sep = ";"
                return pd.read_csv(
                    bio, sep=sep, encoding="utf-8", dtype=str, engine="python"
                )
            return pd.read_excel(bio, engine="openpyxl", dtype=str)
        except Exception as e:
            st.error(f"Erro ao ler arquivo: {e}")
            return pd.DataFrame()

    @staticmethod
    @st.cache_data(
        ttl=600, show_spinner="🔗 Conectando com Google Sheets (lista_ativos)..."
    )
    def buscar_gsheets() -> pd.DataFrame:
        try:
            from streamlit_gsheets import GSheetsConnection  # type: ignore

            conn = st.connection("gsheets", type=GSheetsConnection)
            raw = conn.read(
                spreadsheet=Config.URL_LISTA_ATIVOS,
                worksheet=Config.WORKSHEET_ATIVOS,
            )
            if raw is not None and not raw.empty:
                return DataLoader._processar_lista_ativos(raw)
        except Exception:
            pass
        for url in (
            f"https://docs.google.com/spreadsheets/d/{Config.SHEET_ID_ATIVOS}"
            f"/gviz/tq?tqx=out:csv&sheet={Config.WORKSHEET_ATIVOS}",
            f"https://docs.google.com/spreadsheets/d/{Config.SHEET_ID_ATIVOS}"
            f"/export?format=csv&gid=0",
        ):
            try:
                raw = pd.read_csv(url)
                if raw is not None and not raw.empty:
                    return DataLoader._processar_lista_ativos(raw)
            except Exception as e:
                st.warning(f"⚠️ Falha ao carregar lista_ativos: {e}")
        return pd.DataFrame()

    @staticmethod
    def _processar_lista_ativos(raw: pd.DataFrame) -> pd.DataFrame:
        if raw is None or raw.empty:
            return pd.DataFrame()
        raw.columns = raw.columns.astype(str).str.strip()
        rename_map = {}
        for col in raw.columns:
            col_upper = col.upper().strip()
            if col_upper in ("LOGIN", "MATRÍCULA", "MATRICULA", "ID"):
                rename_map[col] = "Login"
            elif col_upper in ("TÉCNICO", "TECNICO", "NOME", "NOME TÉCNICO"):
                rename_map[col] = "Técnico"
            elif col_upper in ("MONITOR", "GESTOR", "SUPERVISOR"):
                rename_map[col] = "Monitor"
            elif col_upper in ("BASE", "REGIÃO", "REGIAO"):
                rename_map[col] = "Base"
        raw = raw.rename(columns=rename_map)
        cols_uteis = [
            c for c in ["Login", "Técnico", "Monitor", "Base"] if c in raw.columns
        ]
        if "Login" not in cols_uteis:
            return pd.DataFrame()
        raw = raw[cols_uteis].copy()
        raw["Login"] = (
            raw["Login"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
            .str.upper()
        )
        raw = raw[raw["Login"].str.strip() != ""]
        raw = raw[~raw["Login"].isin(["NAN", "NONE", "NULL", "N/A"])]
        return raw.drop_duplicates(subset=["Login"], keep="last").reset_index(drop=True)

    @staticmethod
    @st.cache_data(show_spinner=False)
    def preparar_base(df: pd.DataFrame, df_gs: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame()

        df = df.copy()
        df.columns = df.columns.astype(str).str.strip().str.upper()
        df.attrs["total_importado"] = len(df)

        # 1. Remover Suspensos
        col_atv = detectar_col_status_atividade(df)
        n_susp = 0
        if col_atv:
            serie_atv = df[col_atv].fillna("").astype(str).str.strip().str.upper()
            mask_susp = (
                serie_atv.str.contains("SUSP", na=False)
                | serie_atv.eq("SUSPENSO")
                | serie_atv.eq("SUSPENSA")
            )
            n_susp = int(mask_susp.sum())
            df = df[~mask_susp].copy()
        df.attrs["col_status_atividade"] = col_atv
        df.attrs["removidos_suspensos"] = n_susp

        # 2. Remover Contratos Inválidos
        col_con = detectar_col_contrato(df)
        n_invalidos = 0
        if col_con:
            serie_con = (
                df[col_con]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.upper()
                .str.replace(r"\.0$", "", regex=True)
            )
            mask_invalido = serie_con.isin(VAZIOS_CONTRATO)
            n_invalidos = int(mask_invalido.sum())
            df = df[~mask_invalido].copy()
        df.attrs["col_contrato"] = col_con
        df.attrs["removidos_contrato"] = n_invalidos

        if df.empty:
            st.warning(
                "⚠️ Base ficou vazia após remoção de suspensos e contratos inválidos."
            )
            return pd.DataFrame()

        # 3. Total de Tarefas
        col_tot = Utils.buscar_coluna(df, ["TOTAL DE TAREFAS", "QTD TAREFAS"])
        if col_tot:
            s_num = (
                pd.to_numeric(
                    df[col_tot].astype(str).str.replace(",", "."), errors="coerce"
                )
                .fillna(1)
                .round()
            )
            df["TOTAL DE TAREFAS"] = s_num.astype("int64")
        else:
            df["TOTAL DE TAREFAS"] = pd.Series(1, index=df.index, dtype="int64")

        # 4. Merge com lista_ativos
        col_login = Utils.buscar_coluna(
            df,
            ["LOGIN DO TÉCNICO", "LOGIN DO TECNICO", "LOGIN", "USUÁRIO", "MATRÍCULA"],
        )
        df.attrs["merge_aplicado"] = False
        df.attrs["merge_matches"] = 0
        df.attrs["merge_total"] = len(df)

        if col_login and not df_gs.empty and "Login" in df_gs.columns:
            df[col_login] = (
                df[col_login]
                .astype(str)
                .str.replace(r"\.0$", "", regex=True)
                .str.strip()
                .str.upper()
            )
            df = df.drop(
                columns=[c for c in ["TÉCNICO", "MONITOR", "Base"] if c in df.columns],
                errors="ignore",
            )
            df = df.merge(
                df_gs,
                left_on=col_login,
                right_on="Login",
                how="left",
                suffixes=("", "_gs"),
            )
            if "Login" in df.columns and col_login != "Login":
                df = df.drop(columns=["Login"], errors="ignore")
            if "Técnico" in df.columns:
                df.attrs["merge_matches"] = int(df["Técnico"].notna().sum())
                df.attrs["merge_aplicado"] = True

        if "Técnico" not in df.columns:
            col_tec_orig = Utils.buscar_coluna(
                df, ["TECNICO", "NOME TECNICO", "NOME DO TECNICO", "TÉCNICO"]
            )
            df["Técnico"] = (
                df[col_tec_orig]
                if col_tec_orig and col_tec_orig in df.columns
                else "NÃO MAPEADO"
            )
        if "Monitor" not in df.columns:
            col_mon_orig = Utils.buscar_coluna(
                df, ["MONITOR", "GESTOR", "SUPERVISOR", "NOME MONITOR"]
            )
            df["Monitor"] = (
                df[col_mon_orig]
                if col_mon_orig and col_mon_orig in df.columns
                else "SEM MONITOR"
            )

        df["TÉCNICO"] = (
            df["Técnico"].fillna("NÃO MAPEADO").astype(str).str.strip().str.upper()
        )
        df["MONITOR"] = (
            df["Monitor"].fillna("SEM MONITOR").astype(str).str.strip().str.upper()
        )
        df = df.drop(columns=["Técnico", "Monitor"], errors="ignore")
        df.loc[df["TÉCNICO"].isin(["", "NAN", "NONE", "NULL"]), "TÉCNICO"] = (
            "NÃO MAPEADO"
        )
        df.loc[df["MONITOR"].isin(["", "NAN", "NONE", "NULL"]), "MONITOR"] = (
            "SEM MONITOR"
        )

        # 5. Regiões
        col_cid = Utils.buscar_coluna(df, ["CIDADE", "LOCALIDADE"])
        cidade = (
            df[col_cid].fillna("").astype(str).str.strip().str.upper()
            if col_cid
            else pd.Series("", index=df.index)
        )
        df["REGIÃO"] = np.select(
            [
                cidade.isin(["SAO PAULO"]),
                cidade.isin(
                    [
                        "GUARULHOS",
                        "ARUJA",
                        "MOGI DAS CRUZES",
                        "SUZANO",
                        "ITAQUAQUECETUBA",
                        "FERRAZ DE VASCONCELOS",
                        "POA",
                    ]
                ),
                cidade.isin(
                    [
                        "SANTO ANDRE",
                        "SAO BERNARDO DO CAMPO",
                        "SAO CAETANO DO SUL",
                        "DIADEMA",
                        "MAUA",
                        "RIBEIRAO PIRES",
                        "RIO GRANDE DA SERRA",
                    ]
                ),
            ],
            ["LESTE", "GRU", "ABCDM"],
            default="OUTRAS",
        )

        # 6. Status Contrato
        col_status = Utils.buscar_coluna(
            df, ["STATUS DA O.S 1", "STATUS OS 1", "STATUS CONTRATO"]
        )
        df["Status Contrato"] = (
            Utils.classificar_status(df[col_status]) if col_status else "Pendente"
        )

        # 7. Classificação centralizada
        df, df["TIPO_SERVICO"] = classificar_tipo_servico(df)

        # 8. Motivo de Baixa
        col_cod = Utils.buscar_coluna(
            df, ["CÓD DE BAIXA 1", "COD DE BAIXA 1", "MOTIVO DE BAIXA"]
        )
        df["_COL_BAIXA"] = df[col_cod].astype(str) if col_cod else ""

        # 9. Data Agenda
        col_data = Utils.buscar_coluna(df, ["DATA", "DT AGENDA", "DATA AGENDA"])
        df["_DATA_AGENDA"] = (
            pd.to_datetime(df[col_data], errors="coerce", dayfirst=True)
            if col_data
            else pd.NaT
        )
        return df


# ═══════════════════════════════════════════════════════════════════════
# MOTOR ANALÍTICO (BUSINESS LOGIC)
# ═══════════════════════════════════════════════════════════════════════
class Motor:
    """Cálculos vetorizados de projeções, SLA, causas e rankings."""

    @staticmethod
    def _soma_status(df: pd.DataFrame, status: str) -> float:
        return float(df.loc[df["Status Contrato"] == status, "TOTAL DE TAREFAS"].sum())

    @staticmethod
    def quebra_atual(df: pd.DataFrame) -> Tuple[float, float]:
        if df.empty:
            return 0.0, 0.0
        exe = Motor._soma_status(df, "Executada")
        nex = Motor._soma_status(df, "Não Executada")
        cons = exe + nex
        return cons, (nex / cons) if cons > 0 else 0.0

    @staticmethod
    def projetar(df: pd.DataFrame, p: float) -> Dict[str, float]:
        if df.empty:
            return dict(
                alocado=0.0,
                exec=0.0,
                naoexec=0.0,
                pend=0.0,
                quebra_atual=0.0,
                fechamento_proj=0.0,
                naoexec_proj=0.0,
            )
        aloc = float(df["TOTAL DE TAREFAS"].sum())
        exe = Motor._soma_status(df, "Executada")
        nex = Motor._soma_status(df, "Não Executada")
        pen = max(0.0, aloc - exe - nex)
        _, qa = Motor.quebra_atual(df)
        nex_proj = nex + (pen * p)
        return dict(
            alocado=aloc,
            exec=exe,
            naoexec=nex,
            pend=pen,
            quebra_atual=qa,
            fechamento_proj=(nex_proj / aloc) if aloc > 0 else 0.0,
            naoexec_proj=nex_proj,
        )

    @staticmethod
    def folga_sla(df: pd.DataFrame, sla: float) -> Dict[str, Any]:
        if df.empty:
            return dict(
                alocado=0,
                exec=0,
                naoexec=0,
                pend=0,
                limite_ne_total=0,
                folga_ne_pendente=0,
                folga_pct_pendente=0,
                precisa_executar_pendente=0,
                estourado=False,
            )
        aloc = float(df["TOTAL DE TAREFAS"].sum())
        exe = Motor._soma_status(df, "Executada")
        nex = Motor._soma_status(df, "Não Executada")
        pen = max(0.0, aloc - exe - nex)
        limite = sla * aloc
        folga_tot = limite - nex
        folga_pen = max(0.0, min(pen, folga_tot))
        return dict(
            alocado=aloc,
            exec=exe,
            naoexec=nex,
            pend=pen,
            limite_ne_total=limite,
            folga_ne_pendente=folga_pen,
            folga_pct_pendente=(folga_pen / pen) if pen > 0 else 0,
            precisa_executar_pendente=max(0.0, pen - folga_pen),
            estourado=folga_tot < 0,
        )

    @staticmethod
    def tabela_cenarios(
        df: pd.DataFrame,
        grupo: str,
        p_ot: float,
        p_base: float,
        p_pess: float,
        min_aloc: float = 5,
    ) -> pd.DataFrame:
        if df.empty or grupo not in df.columns:
            return pd.DataFrame()
        pv = pd.pivot_table(
            df,
            index=grupo,
            columns="Status Contrato",
            values="TOTAL DE TAREFAS",
            aggfunc="sum",
            fill_value=0,
        )
        for c in Config.STATUS_ORDEM:
            if c not in pv.columns:
                pv[c] = 0.0
        out = pv.reset_index()
        out["Considerado"] = out["Executada"] + out["Não Executada"]
        out["Alocado"] = out["Considerado"] + out["Pendente"]
        out["Quebra Atual"] = np.where(
            out["Considerado"] > 0, out["Não Executada"] / out["Considerado"], 0
        )
        for nome, p in [("Otimista", p_ot), ("Base", p_base), ("Pessimista", p_pess)]:
            out[f"Fechamento {nome}"] = np.where(
                out["Alocado"] > 0,
                (out["Não Executada"] + out["Pendente"] * p) / out["Alocado"],
                0,
            )
        return out[out["Alocado"] >= min_aloc].sort_values(
            "Fechamento Base", ascending=False
        )

    @staticmethod
    def tecnicos_criticos(
        df: pd.DataFrame,
        segmento: str,
        p_base: float,
        min_aloc: float,
        top_n: int,
        p_ot: float = 0.15,
        p_pess: float = 0.50,
    ) -> pd.DataFrame:
        # Suporte para "TODOS" ou segmento específico
        if segmento and segmento != "TODOS" and "TIPO_SERVICO" in df.columns:
            df_seg = df[df["TIPO_SERVICO"] == segmento].copy()
        else:
            df_seg = df.copy()

        if df_seg.empty:
            return pd.DataFrame()
        tab = Motor.tabela_cenarios(df_seg, "TÉCNICO", p_ot, p_base, p_pess, min_aloc)
        return tab.head(top_n) if not tab.empty else pd.DataFrame()

    @staticmethod
    def _normalizar_baixa(df_nex: pd.DataFrame, col_baixa: str) -> pd.DataFrame:
        df_nex = df_nex.copy()
        df_nex["_baixa_norm"] = (
            df_nex[col_baixa]
            .fillna("Sem Registro")
            .astype(str)
            .str.strip()
            .str.upper()
            .replace({"NAN": "Sem Registro", "": "Sem Registro"})
        )
        return df_nex

    @staticmethod
    def causa_raiz(df: pd.DataFrame, col_baixa: str, top_n: int = 8) -> pd.DataFrame:
        df_nex = df[df["Status Contrato"] == "Não Executada"]
        if df_nex.empty or col_baixa not in df_nex.columns:
            return pd.DataFrame()
        df_nex = Motor._normalizar_baixa(df_nex, col_baixa)
        res = (
            df_nex.groupby("_baixa_norm")["TOTAL DE TAREFAS"]
            .sum()
            .nlargest(top_n)
            .reset_index()
        )
        res.columns = ["Motivo de Baixa", "Volume"]
        total = res["Volume"].sum()
        res["% do Total"] = res["Volume"] / total if total > 0 else 0
        res["Acumulado"] = res["% do Total"].cumsum()
        return res

    @staticmethod
    def causa_por_monitor(
        df: pd.DataFrame, col_baixa: str, top_n_monitores: int = 10
    ) -> pd.DataFrame:
        df_nex = df[df["Status Contrato"] == "Não Executada"]
        if df_nex.empty or col_baixa not in df_nex.columns:
            return pd.DataFrame()
        df_nex = Motor._normalizar_baixa(df_nex, col_baixa)

        vol_por_mon = (
            df_nex.groupby("MONITOR")["TOTAL DE TAREFAS"]
            .sum()
            .nlargest(top_n_monitores)
            .reset_index()
        )
        vol_por_mon.columns = ["Monitor", "Total NE"]

        motivo_top = (
            df_nex.groupby(["MONITOR", "_baixa_norm"])["TOTAL DE TAREFAS"]
            .sum()
            .reset_index()
            .sort_values(["MONITOR", "TOTAL DE TAREFAS"], ascending=[True, False])
            .groupby("MONITOR")
            .first()
            .reset_index()
        )
        motivo_top.columns = ["Monitor", "Motivo Principal", "Vol. Motivo"]
        result = vol_por_mon.merge(motivo_top, on="Monitor", how="left")
        result["% do Motivo"] = np.where(
            result["Total NE"] > 0, result["Vol. Motivo"] / result["Total NE"], 0
        )
        return result

    @staticmethod
    def causa_por_regiao(df: pd.DataFrame, col_baixa: str) -> pd.DataFrame:
        df_nex = df[df["Status Contrato"] == "Não Executada"]
        if df_nex.empty or col_baixa not in df_nex.columns:
            return pd.DataFrame()
        df_nex = Motor._normalizar_baixa(df_nex, col_baixa)
        top_motivos = (
            df_nex.groupby("_baixa_norm")["TOTAL DE TAREFAS"]
            .sum()
            .nlargest(10)
            .index.tolist()
        )
        df_top = df_nex[df_nex["_baixa_norm"].isin(top_motivos)]
        if df_top.empty:
            return pd.DataFrame()
        pivot = (
            pd.pivot_table(
                df_top,
                index="_baixa_norm",
                columns="REGIÃO",
                values="TOTAL DE TAREFAS",
                aggfunc="sum",
                fill_value=0,
            )
            .reset_index()
            .rename(columns={"_baixa_norm": "Motivo"})
        )
        pivot["Total"] = pivot.iloc[:, 1:].sum(axis=1)
        return pivot.sort_values("Total", ascending=False)

    @staticmethod
    def backoffice_fila(df: pd.DataFrame) -> pd.DataFrame:
        df_fila = df[df["Status Contrato"].isin(["Não Executada", "Pendente"])]
        if df_fila.empty:
            return pd.DataFrame()
        agg = (
            df_fila.groupby(["MONITOR", "TÉCNICO", "TIPO_SERVICO", "Status Contrato"])[
                "TOTAL DE TAREFAS"
            ]
            .sum()
            .reset_index()
        )
        pivot = pd.pivot_table(
            agg,
            index=["MONITOR", "TÉCNICO", "TIPO_SERVICO"],
            columns="Status Contrato",
            values="TOTAL DE TAREFAS",
            aggfunc="sum",
            fill_value=0,
        ).reset_index()
        for col in ["Não Executada", "Pendente"]:
            if col not in pivot.columns:
                pivot[col] = 0
        pivot["Total Fila"] = pivot["Não Executada"] + pivot["Pendente"]
        pivot["Prioridade"] = pivot["Não Executada"] * 2 + pivot["Pendente"]
        pivot["Classificação"] = np.select(
            [
                pivot["Prioridade"] >= 20,
                pivot["Prioridade"] >= 10,
                pivot["Prioridade"] >= 5,
            ],
            ["🔴 CRÍTICO", "🟠 ALTA", "🟡 MÉDIA"],
            default="🟢 BAIXA",
        )
        pivot = pivot.sort_values("Prioridade", ascending=False).reset_index(drop=True)
        return pivot[
            [
                "Classificação",
                "MONITOR",
                "TÉCNICO",
                "TIPO_SERVICO",
                "Não Executada",
                "Pendente",
                "Total Fila",
                "Prioridade",
            ]
        ].rename(
            columns={
                "MONITOR": "Monitor",
                "TÉCNICO": "Técnico",
                "TIPO_SERVICO": "Segmento",
            }
        )

    @staticmethod
    def backoffice_reincidencia(
        df: pd.DataFrame, col_baixa: str, min_ocorrencias: int = 2
    ) -> pd.DataFrame:
        df_nex = df[df["Status Contrato"] == "Não Executada"]
        if df_nex.empty or col_baixa not in df_nex.columns:
            return pd.DataFrame()
        df_nex = Motor._normalizar_baixa(df_nex, col_baixa)
        df_nex = df_nex[df_nex["_baixa_norm"] != "SEM REGISTRO"]
        if df_nex.empty:
            return pd.DataFrame()
        agg = (
            df_nex.groupby(["TÉCNICO", "_baixa_norm", "MONITOR"])
            .agg(
                Ocorrencias=("TOTAL DE TAREFAS", "count"),
                Volume=("TOTAL DE TAREFAS", "sum"),
            )
            .reset_index()
        )
        reincidentes = agg[agg["Ocorrencias"] >= min_ocorrencias]
        if reincidentes.empty:
            return pd.DataFrame()
        return (
            reincidentes.sort_values(
                ["Ocorrencias", "Volume"], ascending=[False, False]
            )
            .reset_index(drop=True)
            .rename(
                columns={
                    "TÉCNICO": "Técnico",
                    "_baixa_norm": "Motivo",
                    "MONITOR": "Monitor",
                }
            )[["Técnico", "Motivo", "Ocorrencias", "Volume", "Monitor"]]
        )

    @staticmethod
    def backoffice_ranking_criticos(df: pd.DataFrame, top_n: int = 15) -> pd.DataFrame:
        df_fila = df[df["Status Contrato"].isin(["Não Executada", "Pendente"])]
        if df_fila.empty:
            return pd.DataFrame()
        return (
            df_fila.groupby(["TÉCNICO", "MONITOR"])
            .agg(
                Total_Fila=("TOTAL DE TAREFAS", "sum"),
                Qtd_OS=("TOTAL DE TAREFAS", "count"),
            )
            .reset_index()
            .sort_values("Total_Fila", ascending=False)
            .head(top_n)
            .rename(
                columns={
                    "TÉCNICO": "Técnico",
                    "MONITOR": "Monitor",
                    "Total_Fila": "Total na Fila",
                    "Qtd_OS": "Qtd OS",
                }
            )
        )

    @staticmethod
    def matriz_resumo(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame()
        df_valid = df[df["TIPO_SERVICO"] != "Outros"].copy()
        if df_valid.empty:
            return pd.DataFrame()

        df_valid["_executadas"] = np.where(
            df_valid["Status Contrato"] == "Executada", df_valid["TOTAL DE TAREFAS"], 0
        )
        df_valid["_nao_executadas"] = np.where(
            df_valid["Status Contrato"] == "Não Executada",
            df_valid["TOTAL DE TAREFAS"],
            0,
        )

        grp = (
            df_valid.groupby(["MONITOR", "TIPO_SERVICO"])
            .agg(
                executados=("_executadas", "sum"),
                nao_executados=("_nao_executadas", "sum"),
                total_tarefas=("TOTAL DE TAREFAS", "sum"),
            )
            .reset_index()
        )
        grp["denominador"] = grp["executados"] + grp["nao_executados"]
        grp["pct"] = np.where(
            grp["denominador"] > 0, grp["nao_executados"] / grp["denominador"], 0.0
        )

        pivot = grp.pivot_table(
            index="MONITOR", columns="TIPO_SERVICO", values="pct", fill_value=0.0
        )
        for t in Config.ORDEM_TIPOS:
            if t not in pivot.columns:
                pivot[t] = 0.0
        pivot = pivot[Config.ORDEM_TIPOS]

        exec_tot = df_valid.groupby("MONITOR")["_executadas"].sum()
        ne_tot = df_valid.groupby("MONITOR")["_nao_executadas"].sum()
        tar_tot = df_valid.groupby("MONITOR")["TOTAL DE TAREFAS"].sum()
        df_tot = pd.DataFrame({"exec": exec_tot, "ne": ne_tot, "tar": tar_tot}).fillna(
            0
        )

        pivot["Quebra Geral"] = np.where(
            (df_tot["exec"] + df_tot["ne"]) > 0,
            df_tot["ne"] / (df_tot["exec"] + df_tot["ne"]),
            0.0,
        )
        pivot["Total Tarefas"] = df_tot["tar"].astype(int)
        pivot = pivot.reset_index().rename(columns={"MONITOR": "Monitor"})

        total_row: Dict[str, Any] = {"Monitor": "Total Geral"}
        for tipo in Config.ORDEM_TIPOS:
            sub = df_valid[df_valid["TIPO_SERVICO"] == tipo]
            ex = sub["_executadas"].sum()
            ne = sub["_nao_executadas"].sum()
            total_row[tipo] = ne / (ex + ne) if (ex + ne) > 0 else 0.0

        ex_g = df_valid["_executadas"].sum()
        ne_g = df_valid["_nao_executadas"].sum()
        total_row["Quebra Geral"] = ne_g / (ex_g + ne_g) if (ex_g + ne_g) > 0 else 0.0
        total_row["Total Tarefas"] = int(df_valid["TOTAL DE TAREFAS"].sum())

        return pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)


# ═══════════════════════════════════════════════════════════════════════
# COMPONENTES VISUAIS AVANÇADOS
# ═══════════════════════════════════════════════════════════════════════
def render_dataframe_profundo(
    df: pd.DataFrame,
    titulo: str,
    icone: str,
    color_col: Optional[str] = None,
    meta: float = 0.20,
    height: int = 400,
) -> None:
    st.markdown(
        f'<div style="background:#FFFFFF;border-radius:0.75rem;padding:1rem 1.2rem;'
        f'box-shadow:0 2px 8px rgba(0,0,0,0.05);margin-bottom:0.5rem;">'
        f'<div style="font-size:1rem;font-weight:700;color:#0F172A;'
        f'display:flex;align-items:center;gap:0.5rem;">'
        f"<span>{icone}</span><span>{titulo}</span>"
        f'<span style="font-size:0.68rem;background:#E0F2FE;color:#0369A1;'
        f'padding:0.15rem 0.5rem;border-radius:999px;">{len(df)} registros</span>'
        f"</div></div>",
        unsafe_allow_html=True,
    )
    if df.empty:
        st.info("Sem dados para exibir.")
        return

    df_disp = df.copy()
    _COLS_INT = [
        "Executada",
        "Não Executada",
        "Pendente",
        "Alocado",
        "Considerado",
        "Qtd Não Executadas",
        "Volume",
        "Total NE",
        "Vol. Motivo",
        "Total Fila",
        "Prioridade",
        "Ocorrencias",
        "Total na Fila",
        "Qtd OS",
        "TOTAL DE TAREFAS",
        "Total Tarefas",
        "Qtde. O.S.",
    ]
    for col in _COLS_INT:
        if col in df_disp.columns:
            df_disp[col] = (
                pd.to_numeric(df_disp[col], errors="coerce").fillna(0).astype(int)
            )

    fmt_cols = [
        "Quebra Atual",
        "Fechamento Otimista",
        "Fechamento Base",
        "Fechamento Pessimista",
        "% do Total",
        "Acumulado",
        "% no Segmento",
        "% do Motivo",
    ]
    fmt_dict: dict[str, Any] = {c: "{:.2%}" for c in fmt_cols if c in df_disp.columns}
    for col in _COLS_INT:
        if col in df_disp.columns:
            fmt_dict[col] = "{:,.0f}"

    # Configurar cores condicionais
    condicao_cores = None
    if color_col and color_col in df_disp.columns:
        condicao_cores = {
            "coluna": color_col,
            "meta": meta,
            "acima_meta": {"bg": "#FEE2E2", "text": "#991B1B", "bold": True},
            "perto_meta": {"bg": "#FEF9C3", "text": "#854D0E", "bold": True},
            "abaixo_meta": {"bg": "#DCFCE7", "text": "#166534", "bold": True},
        }

    # Destacar coluna Quebra Atual
    destaque_col = None
    if "Quebra Atual" in df_disp.columns:
        destaque_col = {
            "coluna": "Quebra Atual",
            "bg": "#1E293B",
            "text": "#FFFFFF",
            "bold": True,
        }

    render_table_html(
        df_disp,
        fmt=fmt_dict,
        color_rules=condicao_cores,
        num_cols=list(df_disp.columns),
        height=height,
    )


def estilizar_matriz(df: pd.DataFrame, meta: float):
    """Prepara dados para render_table_html na matriz."""
    cols_pct = [c for c in df.columns if c not in ("Monitor", "Total Tarefas")]

    fmt: Dict[str, Any] = {c: _fmt_pct_br for c in cols_pct}
    if "Total Tarefas" in df.columns:
        fmt["Total Tarefas"] = _fmt_int_br

    # Configurar cores por coluna de porcentagem
    condicoes = {}
    for col in cols_pct:
        condicoes[col] = {
            "meta": meta,
            "acima_meta": {"bg": "#FEE2E2", "text": "#991B1B", "bold": True},
            "abaixo_meta": {"bg": "#D1FAE5", "text": "#065F46", "bold": True},
        }

    # Estilo especial para Total Geral
    linha_total = {"coluna": "Monitor", "valor": "TOTAL GERAL"}

    return df, fmt, condicoes, linha_total


def render_dataframe(
    df: pd.DataFrame,
    titulo: str,
    icone: str,
    fmt: Optional[Dict[str, str]] = None,
    color_col: Optional[str] = None,
    color_meta: float = 0.20,
    color_invertido: bool = False,
    height: int = 400,
) -> None:
    """Wrapper de compatibilidade."""
    render_dataframe_profundo(df, titulo, icone, color_col, color_meta, height)


# ═══════════════════════════════════════════════════════════════════════
# HEROS E HEADERS DINÂMICOS
# ═══════════════════════════════════════════════════════════════════════
def html_resultado_base(regioes: List[str], total: int) -> str:
    badges = "".join(
        [
            f'<span style="padding:0.3rem 0.9rem;border-radius:999px;'
            f"font-size:0.82rem;font-weight:700;border:2px solid;"
            f'background:{CORES_REGIAO.get(r, CORES_REGIAO["OUTRAS"])["bg"]};'
            f'color:{CORES_REGIAO.get(r, CORES_REGIAO["OUTRAS"])["text"]};'
            f'border-color:{CORES_REGIAO.get(r, CORES_REGIAO["OUTRAS"])["border"]};">'
            f"{r}</span>"
            for r in sorted(regioes)
        ]
    )
    total_fmt = f"{total:,}".replace(",", ".")
    return (
        '<div style="background:linear-gradient(135deg, #0F172A 0%, #1E3A5F 100%);'
        "padding:1rem 1.5rem;border-radius:0.75rem;margin-bottom:1.5rem;"
        "display:flex;align-items:center;flex-wrap:wrap;gap:0.6rem;"
        'box-shadow:0 4px 12px rgba(0,0,0,0.15);">'
        '<span style="color:#94A3B8;font-size:0.8rem;font-weight:700;'
        'text-transform:uppercase;letter-spacing:0.08em;">📋 Resultado da Base:</span>'
        f"{badges}"
        '<span style="color:#FFFFFF;font-size:0.78rem;margin-left:auto;'
        f'font-weight:700;">{total_fmt} registros</span>'
        "</div>"
    )


def render_hero_topo_fixo(
    titulo: str,
    subtitulo: str,
    regioes: List[str],
    total: int,
    badge: str = "",
) -> None:
    badge_html = ""
    if badge:
        badge_html = (
            f'<span style="display:inline-block;background:rgba(255,255,255,0.20);'
            f"padding:5px 16px;border-radius:20px;font-size:12px;font-weight:700;"
            f"margin-top:10px;letter-spacing:0.6px;text-transform:uppercase;"
            f'color:white;border:1px solid rgba(255,255,255,0.30);">'
            f"{badge}</span>"
        )
    resultado_html = html_resultado_base(regioes, total) if total > 0 else ""
    st.markdown(
        f'<div style="background:rgba(248,250,252,0.95);padding:0.5rem 0;border-radius:14px;">'
        f'<div style="background:linear-gradient(135deg, #012869 0%, #1E40AF 50%, #F37C04 100%);'
        f"padding:28px 40px;border-radius:14px;color:white;"
        f"box-shadow:0 10px 40px rgba(1,40,105,0.20);margin-bottom:12px;"
        f'position:relative;overflow:hidden;border:1px solid rgba(255,255,255,0.10);">'
        f'<div style="position:absolute;top:50%;right:-100px;transform:translateY(-50%);'
        f"width:420px;height:420px;background:radial-gradient(circle at center,"
        f"rgba(255,180,90,0.35) 0%, rgba(243,124,4,0.20) 35%,"
        f"rgba(232,93,4,0.08) 60%, transparent 78%);"
        f'border-radius:50%;pointer-events:none;filter:blur(2px);"></div>'
        f'<div style="position:relative;z-index:2;">'
        f'<h1 style="margin:0;font-size:30px;font-weight:800;color:white!important;'
        f'letter-spacing:-0.5px;text-shadow:0 2px 4px rgba(0,0,0,0.45);">{titulo}</h1>'
        f'<p style="margin:6px 0 0 0;font-size:14px;opacity:0.95;'
        f'color:#F8FAFC;text-shadow:0 1px 3px rgba(0,0,0,0.40);">{subtitulo}</p>'
        f"{badge_html}</div></div>"
        f"{resultado_html}</div>",
        unsafe_allow_html=True,
    )


def render_hero_upload() -> None:
    st.markdown(
        '<div style="background:linear-gradient(135deg, #012869 0%, #1E40AF 50%, #F37C04 100%);'
        "padding:32px 44px;border-radius:14px;color:white;"
        "box-shadow:0 10px 40px rgba(1,40,105,0.25);margin-bottom:24px;"
        'position:relative;overflow:hidden;border:1px solid rgba(255,255,255,0.10);">'
        '<div style="position:absolute;top:50%;right:-100px;transform:translateY(-50%);'
        "width:420px;height:420px;background:radial-gradient(circle at center,"
        "rgba(255,180,90,0.35) 0%, rgba(243,124,4,0.20) 35%,"
        "rgba(232,93,4,0.08) 60%, transparent 78%);"
        'border-radius:50%;pointer-events:none;filter:blur(2px);"></div>'
        '<div style="position:relative;z-index:2;">'
        '<h1 style="margin:0;font-size:34px;font-weight:800;color:white!important;'
        'letter-spacing:-0.8px;text-shadow:0 2px 4px rgba(0,0,0,0.45);">'
        "📉 Gestão de Quebra de Agenda</h1>"
        '<p style="margin:8px 0 0 0;font-size:15px;opacity:0.95;'
        'color:#F8FAFC;text-shadow:0 1px 3px rgba(0,0,0,0.40);">'
        "Importe a base para gerar o Super Relatório Consolidado</p>"
        '<span style="display:inline-block;background:rgba(255,255,255,0.20);'
        "padding:5px 16px;border-radius:20px;font-size:12px;font-weight:700;"
        "margin-top:12px;letter-spacing:0.6px;text-transform:uppercase;"
        'color:white;border:1px solid rgba(255,255,255,0.30);">SISTEMA TOTALE</span>'
        "</div></div>",
        unsafe_allow_html=True,
    )


# ═══════════════════════════════════════════════════════════════════════
# VISUALIZAÇÃO DOS MÓDULOS DE NAVEGAÇÃO
# ═══════════════════════════════════════════════════════════════════════
def view_resumo_executivo(df: pd.DataFrame, meta_sla: float) -> None:
    render_section("📊 Matriz de Quebra por Monitor e Segmento")

    df_matriz = Motor.matriz_resumo(df)
    if df_matriz.empty:
        st.warning("⚠️ Dados insuficientes para montar a Matriz Executiva.")
        return

    df_proc, fmt, condicoes, linha_total = estilizar_matriz(df_matriz, meta_sla)

    render_table_html(
        df_proc,
        fmt=fmt,
        linha_total=bool(linha_total),
    )

    c1, c2 = st.columns([1, 1])
    with c1:
        st.download_button(
            "📥 Baixar Matriz (Excel)",
            data=Utils.gerar_excel(df_matriz, "Matriz_Resumo"),
            file_name="Matriz_Resumo_Quebra.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def view_analise_detalhada(
    df: pd.DataFrame,
    p_ot: float,
    p_base: float,
    p_pess: float,
    min_aloc: float,
    meta_sla: float,
) -> None:
    tab1, tab2, tab3, tab4 = st.tabs(
        ["📈 Projeções", "🏆 Rankings", "🔍 Causas Raiz", "🛠️ Backoffice"]
    )

    with tab1:
        render_section("📈 Cenários de Projeção de Fechamento")
        tab_monitores = Motor.tabela_cenarios(
            df, "MONITOR", p_ot, p_base, p_pess, min_aloc
        )
        render_dataframe_profundo(
            tab_monitores,
            "Projeção por Monitor",
            "👨‍💼",
            color_col="Fechamento Base",
            meta=meta_sla,
        )

    with tab2:
        render_section("🏆 Técnicos Mais Críticos")
        df_tec = Motor.tecnicos_criticos(
            df, "TODOS", p_base, min_aloc, top_n=20, p_ot=p_ot, p_pess=p_pess
        )
        render_dataframe_profundo(
            df_tec,
            "Top 20 Técnicos com Maior Risco",
            "⚠️",
            color_col="Fechamento Base",
            meta=meta_sla,
        )

    with tab3:
        render_section("🔍 Análise de Causa Raiz (Motivos de Baixa)")
        col_baixa = cast(str, df.attrs.get("_COL_BAIXA", "_COL_BAIXA"))
        df_causa = Motor.causa_raiz(df, col_baixa, top_n=10)
        render_dataframe_profundo(df_causa, "Pareto de Motivos", "🎯")

    with tab4:
        render_section("🛠️ Gestão de Fila e Reincidência (Backoffice)")
        df_fila = Motor.backoffice_fila(df)
        render_dataframe_profundo(df_fila, "Fila Priorizada", "📋")


# ═══════════════════════════════════════════════════════════════════════
# FLUXO PRINCIPAL APLICAÇÃO (STREAMLIT)
# ═══════════════════════════════════════════════════════════════════════
def main() -> None:
    st.sidebar.title("⚙️ Painel de Controle")

    uploaded_file = st.sidebar.file_uploader(
        "Carregar Base de Dados (CSV/XLSX)", type=["csv", "xlsx"]
    )

    if uploaded_file is not None:
        file_bytes = uploaded_file.getvalue()
        raw_df = DataLoader.ler_arquivo(file_bytes, uploaded_file.name)
        df_gs = DataLoader.buscar_gsheets()
        st.session_state["df_memoria"] = DataLoader.preparar_base(raw_df, df_gs)

    df = st.session_state["df_memoria"]

    if df is None or df.empty:
        render_hero_upload()
        st.info("👈 Por favor, faça o upload de uma base de dados para começar.")
        return

    # Filtros na Barra Lateral
    st.sidebar.markdown("---")
    st.sidebar.subheader("🎯 Filtros e Parâmetros")

    regioes_disponiveis = sorted(df["REGIÃO"].unique().tolist())
    regioes_sel = st.sidebar.multiselect(
        "Região", regioes_disponiveis, default=regioes_disponiveis
    )

    df_filtrado = df[df["REGIÃO"].isin(regioes_sel)].copy()

    p_ot = st.sidebar.slider("Probabilidade Otimista (%)", 0, 100, 15, step=5) / 100.0
    p_base = st.sidebar.slider("Probabilidade Base (%)", 0, 100, 30, step=5) / 100.0
    p_pess = (
        st.sidebar.slider("Probabilidade Pessimista (%)", 0, 100, 50, step=5) / 100.0
    )
    min_aloc = float(
        st.sidebar.number_input("Mínimo de Alocações", value=5, min_value=1)
    )

    render_hero_topo_fixo(
        "Super Relatório Corporativo",
        "Análise unificada de desempenho operacional e quebra de agenda",
        regioes_sel,
        len(df_filtrado),
        badge="TOTALE OPERACIONAL",
    )

    aba = st.radio(
        "Navegação",
        ["Resumo Executivo", "Análise Detalhada", "Auditoria"],
        horizontal=True,
    )

    if aba == "Resumo Executivo":
        view_resumo_executivo(df_filtrado, Config.SLA_QUEBRA_MAXIMA)
    elif aba == "Análise Detalhada":
        view_analise_detalhada(
            df_filtrado, p_ot, p_base, p_pess, min_aloc, Config.SLA_QUEBRA_MAXIMA
        )
    elif aba == "Auditoria":
        render_painel_criterios(df_filtrado)
        render_debug_criterios(df_filtrado)


if __name__ == "__main__":
    main()
